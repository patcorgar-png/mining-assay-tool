"""
Mining Assay Data Processing Pipeline
======================================
Loads messy lab Excel assay files, cleans and standardises data,
computes AuEq grades, finds best mineralised intervals, and writes
a professional multi-sheet Excel workbook.

Entry point
-----------
    run_pipeline(filepath, prices, cutoff, output_path)

prices : dict  e.g. {"Au": 1950, "Ag": 24, "Pb": 0.95, "Zn": 1.20}
            Precious metals (Au, Ag) : USD / troy oz
            Base metals (Pb, Zn, Cu): USD / lb
cutoff : float  minimum AuEq g/t for a qualifying interval
"""

from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# ---------------------------------------------------------------------------
# 1. Element / column-name catalogue
# ---------------------------------------------------------------------------

_ELEMENT_ALIASES: dict[str, str] = {
    "gold":      "Au", "au":   "Au",
    "silver":    "Ag", "ag":   "Ag",
    "lead":      "Pb", "pb":   "Pb",
    "zinc":      "Zn", "zn":   "Zn",
    "copper":    "Cu", "cu":   "Cu",
    "iron":      "Fe", "fe":   "Fe",
    "arsenic":   "As", "as":   "As",
    "antimony":  "Sb", "sb":   "Sb",
    "bismuth":   "Bi", "bi":   "Bi",
    "cadmium":   "Cd", "cd":   "Cd",
    "cobalt":    "Co", "co":   "Co",
    "chromium":  "Cr", "cr":   "Cr",
    "manganese": "Mn", "mn":   "Mn",
    "molybdenum":"Mo", "mo":   "Mo",
    "nickel":    "Ni", "ni":   "Ni",
    "phosphorus":"P",  "p":    "P",
    "sulfur":    "S",  "s":    "S",
    "tin":       "Sn", "sn":   "Sn",
    "tungsten":  "W",  "w":    "W",
    "vanadium":  "V",  "v":    "V",
    "barium":    "Ba", "ba":   "Ba",
    "calcium":   "Ca", "ca":   "Ca",
    "potassium": "K",  "k":    "K",
    "magnesium": "Mg", "mg":   "Mg",
    "sodium":    "Na", "na":   "Na",
    "silicon":   "Si", "si":   "Si",
    "titanium":  "Ti", "ti":   "Ti",
}

# Priced in USD/troy oz; all others assumed USD/lb (base metals in %)
_GT_ELEMENTS = {"Au", "Ag", "Pt", "Pd"}

_UNIT_CONVERSIONS: dict[str, float] = {
    "oz/t":  31.1035,
    "opt":   31.1035,
    "ppm":   0.0001,    # ppm -> % for base metals
    "ppb":   0.0000001,
    "g/t":   1.0,
    "g/mt":  1.0,       # g/metric tonne = g/t
    "%":     1.0,
}


# ---------------------------------------------------------------------------
# 2. Header / data-row auto-detection
# ---------------------------------------------------------------------------

_HEADER_RE = re.compile(
    r"\b(au|ag|pb|zn|cu|fe|sample|hole|from|to|depth|interval|assay)\b",
    re.IGNORECASE,
)


def _score_row_as_header(row: pd.Series) -> int:
    score = 0
    for val in row.dropna():
        txt = str(val).strip()
        if _HEADER_RE.search(txt):
            score += 2
        if len(txt) <= 6 and txt.isalpha():
            score += 1
    return score


def detect_header_row(raw: pd.DataFrame) -> tuple[int, int]:
    """
    Scan the first 40 rows to find the header row and first data row.
    Returns (header_row_index, data_start_row_index) -- zero-based.
    """
    best_score, best_idx = -1, 0
    for i in range(min(40, len(raw))):
        score = _score_row_as_header(raw.iloc[i])
        if score > best_score:
            best_score, best_idx = score, i
    data_start = best_idx + 1
    while data_start < len(raw) and raw.iloc[data_start].isna().all():
        data_start += 1
    return best_idx, data_start


# ---------------------------------------------------------------------------
# 3. Load Excel file
# ---------------------------------------------------------------------------

_EXTRA_NA = ["", " ", "N/A", "n/a", "NA", "na", "#N/A", "-", "--",
             "nd", "ND", "bdl", "BDL", "<dl>", "nil", "Nil", "NULL", "null", "none", "None"]


def load_raw(
    filepath: str | Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    sheet: int | str = 0,
) -> pd.DataFrame:
    """
    Load an assay Excel file into a clean DataFrame.

    Parameters
    ----------
    filepath       : path to .xlsx / .xls file
    header_row     : zero-based row index of the header; None => auto-detect
    data_start_row : zero-based row index of first data row; None => auto-detect
    sheet          : sheet index or name (default 0)
    """
    filepath = Path(filepath)
    raw = pd.read_excel(
        filepath, sheet_name=sheet,
        header=None, dtype=object, keep_default_na=False,
        na_values=_EXTRA_NA,
    )
    # Convert all cells to str or NaN — never leave raw numeric/date objects
    raw = raw.map(lambda v: str(v).strip() if pd.notna(v) and str(v).strip() not in _EXTRA_NA else np.nan)
    # Belt-and-suspenders: replace any remaining empty / whitespace-only strings
    raw = raw.replace("", np.nan)
    raw = raw.replace(r"^\s+$", np.nan, regex=True)

    if header_row is None:
        header_row, data_start_auto = detect_header_row(raw)
        if data_start_row is None:
            data_start_row = data_start_auto
    elif data_start_row is None:
        data_start_row = header_row + 1

    col_names = [
        str(v).strip() if pd.notna(v) else f"_col{i}"
        for i, v in enumerate(raw.iloc[header_row])
    ]
    df = raw.iloc[data_start_row:].copy()
    df.columns = col_names
    df = df.loc[:, ~df.columns.str.startswith("_col")]
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


# ---------------------------------------------------------------------------
# 4. Detection-limit cleaning
# ---------------------------------------------------------------------------

_DL_LESS_THAN    = re.compile(r"^[<<]\s*([\d.]+)\s*$")
_DL_GREATER_THAN = re.compile(r"^[>>]\s*([\d.]+)\s*$")


def clean_detection_limits(value):
    """
    '<0.005'  -> 0.0
    '>10000'  -> 10000.0
    Anything else is coerced to float or returned as-is.
    """
    # Handle None / NaN / already-numeric values
    if value is None:
        return np.nan
    if pd.isna(value) if not isinstance(value, (list, dict)) else False:
        return np.nan
    if isinstance(value, (int, float)):
        return float(value) if not np.isnan(float(value)) else np.nan
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s or s in _EXTRA_NA:
        return np.nan
    m = _DL_LESS_THAN.match(s)
    if m:
        return 0.0
    m = _DL_GREATER_THAN.match(s)
    if m:
        return float(m.group(1))
    try:
        return float(s)
    except ValueError:
        return s


# ---------------------------------------------------------------------------
# 5. Unit metadata extraction
# ---------------------------------------------------------------------------

_UNIT_ALIASES = {u.lower(): u.lower() for u in _UNIT_CONVERSIONS}

# Matches units in parentheses like "(ppm)" or "(g/t)", OR as a suffix
# after space/underscore/dash/end-of-string: "Pb ppm", "Pb_ppm", "Au_g/t"
_UNIT_RE = re.compile(
    r"[\(\s_\-](" + "|".join(re.escape(u) for u in _UNIT_CONVERSIONS) + r")[\)\s]?$",
    re.IGNORECASE,
)


def extract_units_from_columns(columns: list[str]) -> dict[str, str]:
    """
    Parse units embedded in column headers.
    Handles: 'Au (g/t)', 'Pb_ppm', 'Zn ppm', 'Ag (oz/t)'.
    """
    result = {}
    for col in columns:
        m = _UNIT_RE.search(col)
        if m:
            result[col] = m.group(1).lower()
    return result


# ---------------------------------------------------------------------------
# 6. Column-name standardisation
# ---------------------------------------------------------------------------

def _parse_col_to_element(col: str) -> Optional[str]:
    base = _UNIT_RE.sub("", col).strip().lower()
    base = re.sub(r"[^a-z]", "", base)
    return _ELEMENT_ALIASES.get(base)


def standardise_columns(df: pd.DataFrame):
    """
    Rename messy assay columns to canonical element symbols.

    Returns (df, col_map, raw_units)
    """
    raw_units = extract_units_from_columns(df.columns.tolist())
    col_map: dict[str, str] = {}
    seen: dict[str, int] = {}

    for col in df.columns:
        elem = _parse_col_to_element(col)
        if elem:
            seen[elem] = seen.get(elem, 0) + 1
            col_map[col] = elem if seen[elem] == 1 else f"{elem}_{seen[elem]}"
        else:
            col_map[col] = col

    df = df.rename(columns=col_map)
    return df, col_map, raw_units


# ---------------------------------------------------------------------------
# 7. Unit conversion
# ---------------------------------------------------------------------------

def apply_unit_conversions(
    df: pd.DataFrame,
    raw_units: dict[str, str],
    col_map: dict[str, str],
) -> pd.DataFrame:
    """
    Convert element columns to canonical units:
      - Au, Ag  -> g/t
      - others  -> %
    """
    df = df.copy()
    for old_col, unit in raw_units.items():
        new_col = col_map.get(old_col, old_col)
        if new_col not in df.columns:
            continue
        elem = _parse_col_to_element(old_col)
        if not elem:
            continue

        u = unit.lower()
        if elem in _GT_ELEMENTS:
            if u in ("oz/t", "opt"):
                factor = 31.1035
            elif u == "ppm":
                factor = 0.001  # ppm -> g/t (1 ppm = 0.001 g/t)
            else:
                factor = 1.0   # already g/t
        else:
            if u == "ppm":
                factor = 0.0001  # ppm -> %
            elif u == "ppb":
                factor = 0.0000001
            else:
                factor = 1.0   # already %

        series = pd.to_numeric(df[new_col], errors="coerce")
        df[new_col] = series * factor if factor != 1.0 else series

    return df


# ---------------------------------------------------------------------------
# 8. Parse hole ID + depth strings
# ---------------------------------------------------------------------------

_HOLEID_DEPTH_RE  = re.compile(
    r"^([A-Za-z]{1,6}[-_]?\d+)\s+(\d+\.?\d*)\s*[-–_]\s*(\d+\.?\d*)$"
)
_HOLEID_DEPTH_RE2 = re.compile(
    r"^([A-Za-z]{1,6}[-_]?\d+)[-_](\d+\.?\d*)[-_](\d+\.?\d*)$"
)
# Lab sample ID: letters-only prefix + space + from_depth (3+ digits) + dash + to_suffix (2-4 digits)
# e.g. "PRC 15335-340" -> hole=PRC, from=15335, to=15340
_HOLEID_DEPTH_RE3 = re.compile(
    r"^([A-Za-z]+(?:\s+[A-Za-z]+)*)\s+(\d{3,})-(\d{2,4})$"
)
# Lab sample ID: letters + space + digits (hole number) + space + from-to
# e.g. "PRC 23 500-505" -> hole="PRC 23", from=500, to=505
_HOLEID_DEPTH_RE4 = re.compile(
    r"^([A-Za-z]+\s+\d+)\s+(\d{2,})-(\d{2,4})$"
)


def _expand_to_depth(from_val: str, to_suffix: str) -> float:
    """Expand abbreviated To depth using From prefix.
    '15335', '340' → 15340.0  (prefix '15' + '340')
    """
    ns = len(to_suffix)
    if len(from_val) > ns:
        return float(from_val[:-ns] + to_suffix)
    return float(to_suffix)


def _try_parse_combined(val: str):
    s = str(val).strip()
    for pat in (_HOLEID_DEPTH_RE, _HOLEID_DEPTH_RE2):
        m = pat.match(s)
        if m:
            return m.group(1), float(m.group(2)), float(m.group(3))
    # Lab sample ID: "PRC 15335-340" -> hole=PRC, from=15335, to=15340
    m = _HOLEID_DEPTH_RE3.match(s)
    if m:
        from_depth = m.group(2)
        to_depth = _expand_to_depth(from_depth, m.group(3))
        from_val = float(from_depth)
        if to_depth > from_val:
            return m.group(1).strip(), from_val, to_depth
    # Lab sample ID: "PRC 23 500-505" -> hole="PRC 23", from=500, to=505
    m = _HOLEID_DEPTH_RE4.match(s)
    if m:
        from_depth = m.group(2)
        to_depth = _expand_to_depth(from_depth, m.group(3))
        from_val = float(from_depth)
        if to_depth > from_val:
            return m.group(1).strip(), from_val, to_depth
    return None


def _find_col(df: pd.DataFrame, patterns: list[str]) -> Optional[str]:
    for pat in patterns:
        for col in df.columns:
            if re.search(pat, col, re.IGNORECASE):
                return col
    return None


def _cols_look_like_depth_interval(from_s: pd.Series, to_s: pd.Series) -> bool:
    """
    Sanity-check: in a valid depth interval, To > From for the majority of rows
    and values should be non-negative and not unreasonably small (> 0.5 ft typical min).
    """
    f = pd.to_numeric(from_s, errors="coerce").dropna()
    t = pd.to_numeric(to_s,   errors="coerce").dropna()
    if len(f) == 0 or len(t) == 0:
        return False
    # Most To values must exceed corresponding From values
    common_idx = f.index.intersection(t.index)
    if len(common_idx) == 0:
        return False
    valid = (t.loc[common_idx] > f.loc[common_idx]).sum()
    frac_valid = valid / len(common_idx)
    # Also check that median To value is reasonably large (> 0.5 — rules out grade columns)
    median_to = t.median()
    return frac_valid >= 0.6 and median_to >= 0.5


def parse_hole_depth_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure df has Hole_ID, From, To columns."""
    df = df.copy()

    hole_col = _find_col(df, [r"hole", r"bhid", r"dh.?id", r"drill", r"holeid", r"^bh$", r"^dh$"])
    from_col = _find_col(df, [r"^from$", r"from_", r"_from", r"depth.?from", r"^start$"])
    to_col   = _find_col(df, [r"^to$",   r"to_",   r"_to",   r"depth.?to",   r"^end$"])

    if from_col and to_col and _cols_look_like_depth_interval(df[from_col], df[to_col]):
        rename = {}
        if hole_col and hole_col != "Hole_ID": rename[hole_col] = "Hole_ID"
        if from_col != "From":    rename[from_col] = "From"
        if to_col   != "To":      rename[to_col]   = "To"
        df.rename(columns=rename, inplace=True)
        if "Hole_ID" not in df.columns:
            df.insert(0, "Hole_ID", "UNKNOWN")
        df["From"] = pd.to_numeric(df["From"], errors="coerce")
        df["To"]   = pd.to_numeric(df["To"],   errors="coerce")
        return df

    # Try combined column (e.g. "BRS009 0-1")
    for col in df.columns:
        sample = df[col].dropna().head(10)
        parsed = [_try_parse_combined(v) for v in sample]
        parsed = [p for p in parsed if p]
        if len(parsed) >= max(3, len(sample) // 2):
            extracted = df[col].apply(
                lambda v: _try_parse_combined(str(v)) if pd.notna(v) else None
            )
            df["Hole_ID"] = extracted.apply(lambda x: x[0] if x else np.nan)
            df["From"]    = extracted.apply(lambda x: float(x[1]) if x else np.nan)
            df["To"]      = extracted.apply(lambda x: float(x[2]) if x else np.nan)
            df.drop(columns=[col], inplace=True)
            return df

    if hole_col:
        df.rename(columns={hole_col: "Hole_ID"}, inplace=True)
    if "Hole_ID" not in df.columns:
        df.insert(0, "Hole_ID", "UNKNOWN")

    # Fallback: find two numeric columns that look like depth intervals
    # Only grab columns whose values are plausibly depths (median > 0.5)
    num_cols = [
        c for c in df.columns
        if c not in ("Hole_ID", "From", "To")
        and pd.to_numeric(df[c], errors="coerce").notna().sum() > len(df) * 0.5
    ]
    # Try every pair in order until we find one that passes the sanity check
    if "From" not in df.columns:
        assigned = False
        for i in range(len(num_cols) - 1):
            for j in range(i + 1, len(num_cols)):
                ca, cb = num_cols[i], num_cols[j]
                if _cols_look_like_depth_interval(df[ca], df[cb]):
                    df.rename(columns={ca: "From", cb: "To"}, inplace=True)
                    assigned = True
                    break
            if assigned:
                break
        # Last resort: just take first two (old behaviour) but warn implicitly
        if not assigned and len(num_cols) >= 2:
            df.rename(columns={num_cols[0]: "From", num_cols[1]: "To"}, inplace=True)

    df["From"] = pd.to_numeric(df.get("From", pd.Series(dtype=float)), errors="coerce")
    df["To"]   = pd.to_numeric(df.get("To",   pd.Series(dtype=float)), errors="coerce")
    return df


# ---------------------------------------------------------------------------
# 9. Full data cleaning pass
# ---------------------------------------------------------------------------

def clean_dataframe(df: pd.DataFrame):
    """
    Run the full cleaning pipeline on a raw DataFrame.
    Returns (cleaned_df, col_map).
    """
    # Extra safety: flush any remaining empty/whitespace strings to NaN before map
    df = df.replace("", np.nan)
    df = df.replace(r"^\s+$", np.nan, regex=True)

    df = df.map(clean_detection_limits)

    # Flush again after map in case any '' slipped through
    df = df.replace("", np.nan)

    df, col_map, raw_units = standardise_columns(df)
    df = apply_unit_conversions(df, raw_units, col_map)
    df = parse_hole_depth_columns(df)

    # Force ALL element symbol columns to numeric (coerce bad values to NaN)
    for col in df.columns:
        if col in set(_ELEMENT_ALIASES.values()):
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Also force From / To to numeric
    for col in ("From", "To"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df[~(df["From"].isna() & df["To"].isna())].reset_index(drop=True)
    return df, col_map


# ---------------------------------------------------------------------------
# 10. AuEq calculation (Python-side, for interval finding)
# ---------------------------------------------------------------------------
#
# Pricing conventions
# -------------------
# Precious metals (Au, Ag) : USD / troy oz
#   AuEq g/t contribution  = grade_g/t * price_$/oz / au_price_$/oz
#
# Base metals (Pb, Zn, Cu) : USD / lb
#   1% grade = 10,000 g/tonne = 10,000 / 453.592 lb/tonne = 22.046 lb/t
#   $/tonne  = grade_% * 22.046 * price_$/lb * recovery
#   AuEq g/t = $/tonne / (au_price_$/oz / 31.1035 g/oz)
#            = grade_% * 22.046 * 31.1035 * price_$/lb * recovery / au_price
#            = grade_% * 685.94 * price_$/lb * recovery / au_price
#
_LB_SCALE = 10_000 * 31.1035 / 453.592  # ~685.94 -- lb/t scaling constant


def compute_aueq_series(
    df: pd.DataFrame,
    prices: dict[str, float],
    recovery: float = 0.75,
) -> pd.Series:
    """Compute AuEq (g/t) for every sample row using prices dict."""
    au_price = prices.get("Au", 1) or 1
    aueq = pd.Series(0.0, index=df.index)

    for elem, price in prices.items():
        if elem not in df.columns:
            continue
        grade = pd.to_numeric(df[elem], errors="coerce").fillna(0)
        if elem in _GT_ELEMENTS:
            aueq = aueq + grade * price / au_price
        else:
            aueq = aueq + grade * _LB_SCALE * price * recovery / au_price

    return aueq


# ---------------------------------------------------------------------------
# 11. Sliding-window best-interval finder
# ---------------------------------------------------------------------------

def find_best_intervals(
    df: pd.DataFrame,
    prices: dict[str, float],
    cutoff: float,
    drop_tolerance: float = 0.5,
    min_length_ft: float = 10.0,
    recovery: float = 0.75,
) -> pd.DataFrame:
    """
    Find the best mineralised intervals per hole using a
    peak-comparison weighted-average sliding window.

    Parameters
    ----------
    df            : cleaned DataFrame with Hole_ID, From, To, element columns
    prices        : metal price dict used for AuEq
    cutoff        : minimum AuEq g/t to qualify
    drop_tolerance: fraction of cutoff that can fall below before breaking
    min_length_ft : minimum interval length in feet
    recovery      : metallurgical recovery for base metals

    Returns
    -------
    DataFrame with qualifying intervals + "No significant intersections" rows.
    """
    df = df.copy()
    df["_AuEq"] = compute_aueq_series(df, prices, recovery=recovery)

    results: list[dict] = []
    elem_cols = [e for e in prices if e in df.columns]
    lower_cut = cutoff * (1.0 - drop_tolerance)

    # $/tonne value scaling
    lb_val = 10_000 / 453.592  # ~22.046 lb per tonne per 1%

    for hole_id, hole_df in df.groupby("Hole_ID", sort=False):
        hole_df = hole_df.sort_values("From").reset_index(drop=True)
        n = len(hole_df)
        best: Optional[dict] = None
        best_aueq = -np.inf

        i = 0
        while i < n:
            if hole_df.at[i, "_AuEq"] < lower_cut:
                i += 1
                continue

            j = i
            while j < n and hole_df.at[j, "_AuEq"] >= lower_cut:
                j += 1

            window = hole_df.iloc[i:j]
            length_ft = float(window["To"].max() - window["From"].min())
            if length_ft < min_length_ft:
                i = j
                continue

            weights = (window["To"] - window["From"]).clip(lower=0)
            if weights.sum() == 0:
                i = j
                continue

            wav_aueq = float(np.average(window["_AuEq"], weights=weights))
            if wav_aueq >= cutoff and wav_aueq > best_aueq:
                best_aueq = wav_aueq
                best = {"window": window, "AuEq_avg": wav_aueq}
            i = j

        if best is None:
            results.append({
                "Hole_ID": hole_id, "From": None, "To": None,
                "Length_ft": None, "Length_m": None,
                "AuEq_avg": None, "Value_per_tonne": None,
                "no_intersection": True,
                **{e: None for e in elem_cols},
            })
        else:
            w = best["window"]
            length_ft = float(w["To"].max() - w["From"].min())
            weights   = (w["To"] - w["From"]).clip(lower=0)

            row: dict = {
                "Hole_ID":        hole_id,
                "From":           float(w["From"].min()),
                "To":             float(w["To"].max()),
                "Length_ft":      round(length_ft, 2),
                "Length_m":       round(length_ft * 0.3048, 2),
                "AuEq_avg":       round(best["AuEq_avg"], 4),
                "no_intersection": False,
            }

            value_pt = 0.0
            for elem in elem_cols:
                g = float(np.average(
                    pd.to_numeric(w[elem], errors="coerce").fillna(0),
                    weights=weights
                )) if elem in w.columns else 0.0
                row[elem] = round(g, 4)
                price = prices.get(elem, 0)
                if elem in _GT_ELEMENTS:
                    value_pt += g * price / 31.1035
                else:
                    value_pt += g * lb_val * price * recovery

            row["Value_per_tonne"] = round(value_pt, 2)
            results.append(row)

    return pd.DataFrame(results) if results else pd.DataFrame()


# ---------------------------------------------------------------------------
# 11a. Per-sample value, flags, and interval variant helpers
# ---------------------------------------------------------------------------

# Individual-metal cutoffs used for the flag column (●/★/✔).
# These match the BRS010 example (Au=0.1 g/t, Ag=12 g/t, Pb=1%, Zn=0.8%).
_DEFAULT_METAL_CUTOFFS: dict[str, float] = {
    "Au": 0.10,   # g/t
    "Ag": 12.0,   # g/t
    "Pb": 1.00,   # %
    "Zn": 0.80,   # %
}


def compute_sample_value(
    df: pd.DataFrame,
    prices: dict[str, float],
    recovery: float = 0.75,
) -> pd.Series:
    """Gross $/tonne value per sample row."""
    LB = 10_000 / 453.592   # ~22.046 lb/t per 1%
    value = pd.Series(0.0, index=df.index)
    for elem, price in prices.items():
        if elem not in df.columns:
            continue
        g = pd.to_numeric(df[elem], errors="coerce").fillna(0.0)
        if elem in _GT_ELEMENTS:
            value += g * price / 31.1035
        else:
            value += g * LB * price * recovery
    return value


def compute_sample_flags(
    df: pd.DataFrame,
    metal_cutoffs: Optional[dict[str, float]] = None,
) -> pd.Series:
    """
    Per-sample flag: '✔' (3+ cutoffs met), '★' (2), '●' (1), '' (0).
    """
    if metal_cutoffs is None:
        metal_cutoffs = _DEFAULT_METAL_CUTOFFS
    counts = pd.Series(0, index=df.index)
    for elem, cut in metal_cutoffs.items():
        if elem in df.columns:
            g = pd.to_numeric(df[elem], errors="coerce").fillna(0.0)
            counts += (g >= cut).astype(int)
    def _sym(n: int) -> str:
        if n >= 3: return "✔"
        if n == 2: return "★"
        if n == 1: return "●"
        return ""
    return counts.map(_sym)


def _trim_samples(
    samples: pd.DataFrame,
    prices: dict[str, float],
    recovery: float = 0.75,
) -> pd.DataFrame:
    """Strip lowest-value end samples to improve average $/t."""
    s    = samples.reset_index(drop=True).copy()
    vals = compute_sample_value(s, prices, recovery).reset_index(drop=True)
    changed = True
    while len(s) > 2 and changed:
        changed = False
        avg  = float(vals.mean())
        head = float(vals.iloc[0])
        tail = float(vals.iloc[-1])
        if head < avg and head <= tail:
            s, vals = s.iloc[1:].reset_index(drop=True), vals.iloc[1:].reset_index(drop=True)
            changed = True
        elif tail < avg:
            s, vals = s.iloc[:-1].reset_index(drop=True), vals.iloc[:-1].reset_index(drop=True)
            changed = True
    return s


def _find_core_samples(
    samples: pd.DataFrame,
    prices: dict[str, float],
    recovery: float = 0.75,
) -> pd.DataFrame:
    """Consecutive sub-interval with the highest average $/t (min 2 samples)."""
    s = samples.reset_index(drop=True)
    n = len(s)
    if n <= 2:
        return s
    vals = compute_sample_value(s, prices, recovery)
    best_avg, best_slice = -1.0, (0, min(2, n))
    for i in range(n):
        for j in range(i + 2, n + 1):
            avg = float(vals.iloc[i:j].mean())
            if avg > best_avg:
                best_avg, best_slice = avg, (i, j)
    return s.iloc[best_slice[0]:best_slice[1]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# 12. Excel style constants
# ---------------------------------------------------------------------------

_THIN        = Side(style="thin", color="CCCCCC")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_DARK_FILL   = PatternFill("solid", fgColor="2F4F4F")
_BLUE_FILL   = PatternFill("solid", fgColor="D9E1F2")
_PINK_FILL   = PatternFill("solid", fgColor="FFE0E0")
_GOLD_FILL   = PatternFill("solid", fgColor="FFF2CC")
_BOLD_WHITE  = Font(bold=True, color="FFFFFF")
_BOLD        = Font(bold=True)
_BOLD_RED    = Font(bold=True, color="C00000")
_GREY_FONT   = Font(color="808080", italic=True)
_FMT_2DP     = "0.00"
_FMT_4DP     = "0.0000"
_FMT_MONEY   = "#,##0.00"

_R = 18   # column R (right analysis block starts here)


# ---------------------------------------------------------------------------
# 13. Prices & Cutoffs block (right panel, starts at col _R)
# ---------------------------------------------------------------------------

def _write_prices_block(
    ws,
    prices: dict[str, float],
    cutoff: float,
    metal_cutoffs: dict[str, float],
    recovery: float,
    start_row: int = 1,
    start_col: int = _R,
) -> int:
    """Write Prices & Cutoffs table. Returns the next available row."""
    sc = start_col
    r  = start_row

    _PRICE_UNITS   = {e: "$/troy oz" if e in _GT_ELEMENTS else "$/lb" for e in prices}
    _CUTOFF_LABELS = {"Au": "g/t", "Ag": "g/t", "Pb": "%", "Zn": "%"}

    # Title banner
    ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc + 3)
    c = ws.cell(row=r, column=sc, value="PRICES & CUTOFFS")
    c.font = _BOLD_WHITE
    c.fill = _DARK_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # Column headers
    for ci, h in enumerate(["Metal", "Price", "Unit", "Cut-off"], sc):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font   = _BOLD
        cell.fill   = _BLUE_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    r += 1

    for elem, price in prices.items():
        ws.cell(row=r, column=sc,     value=elem).border = _THIN_BORDER
        c2 = ws.cell(row=r, column=sc + 1, value=price)
        c2.number_format = "#,##0.00"
        c2.border = _THIN_BORDER
        ws.cell(row=r, column=sc + 2, value=_PRICE_UNITS.get(elem, "")).border = _THIN_BORDER
        cut = metal_cutoffs.get(elem)
        unit = _CUTOFF_LABELS.get(elem, "")
        ws.cell(row=r, column=sc + 3,
                value=f"{cut} {unit}" if cut is not None else "").border = _THIN_BORDER
        r += 1

    # Recovery row
    ws.cell(row=r, column=sc,     value="Recovery").border    = _THIN_BORDER
    cr = ws.cell(row=r, column=sc + 1, value=recovery)
    cr.number_format = "0%"
    cr.border = _THIN_BORDER
    ws.cell(row=r, column=sc + 2, value="").border = _THIN_BORDER
    ws.cell(row=r, column=sc + 3, value="").border = _THIN_BORDER
    r += 1

    # AuEq cutoff row
    ws.cell(row=r, column=sc,     value="AuEq Cutoff").border = _THIN_BORDER
    cc = ws.cell(row=r, column=sc + 1, value=cutoff)
    cc.number_format = _FMT_2DP
    cc.border = _THIN_BORDER
    ws.cell(row=r, column=sc + 2, value="g/t AuEq").border = _THIN_BORDER
    ws.cell(row=r, column=sc + 3, value="").border = _THIN_BORDER
    r += 1

    return r + 1  # leave a blank row gap


# ---------------------------------------------------------------------------
# 14. Best Intervals summary block
# ---------------------------------------------------------------------------

def _write_summary_block(
    ws,
    cleaned_df: pd.DataFrame,
    intervals: pd.DataFrame,
    prices: dict[str, float],
    recovery: float,
    start_row: int,
    start_col: int = _R,
) -> int:
    """Write per-hole Best Intervals summary (Full / Trim / Core). Returns next row."""
    sc = start_col
    r  = start_row

    # Title
    ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc + 8)
    c = ws.cell(row=r, column=sc, value="BEST INTERVALS SUMMARY")
    c.font = _BOLD_WHITE
    c.fill = _DARK_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # Column headers
    hdrs = [
        "Hole ID",
        "Full: From", "Full: To", "Full $/t",
        "Trim: From", "Trim: To", "Trim $/t",
        "Core: From", "Core: To", "Core $/t",
    ]
    for ci, h in enumerate(hdrs, sc):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font   = _BOLD
        cell.fill   = _BLUE_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1

    for _, irow in intervals.iterrows():
        hole = irow["Hole_ID"]
        is_ni = bool(irow.get("no_intersection", False))
        ws.cell(row=r, column=sc, value=str(hole)).border = _THIN_BORDER

        if is_ni:
            ws.merge_cells(start_row=r, start_column=sc + 1, end_row=r, end_column=sc + 9)
            mc = ws.cell(row=r, column=sc + 1, value="No significant intersections")
            mc.font = _GREY_FONT
            mc.alignment = Alignment(horizontal="center")
            for ci in range(sc, sc + 10):
                ws.cell(row=r, column=ci).border = _THIN_BORDER
        else:
            # Grab the Full-interval samples from cleaned_df
            full_mask = (
                (cleaned_df["Hole_ID"] == hole) &
                (pd.to_numeric(cleaned_df["From"], errors="coerce") >= irow["From"]) &
                (pd.to_numeric(cleaned_df["To"],   errors="coerce") <= irow["To"])
            )
            full_samp = cleaned_df[full_mask].sort_values("From").reset_index(drop=True)

            def _interval_stats(samp):
                if samp.empty:
                    return None, None, None
                frm = float(samp["From"].min())
                to  = float(samp["To"].max())
                val = float(compute_sample_value(samp, prices, recovery).mean())
                return frm, to, val

            full_f, full_t, full_v   = _interval_stats(full_samp)
            trim_samp  = _trim_samples(full_samp,  prices, recovery) if not full_samp.empty else full_samp
            core_samp  = _find_core_samples(full_samp, prices, recovery) if not full_samp.empty else full_samp
            trim_f, trim_t, trim_v  = _interval_stats(trim_samp)
            core_f, core_t, core_v  = _interval_stats(core_samp)

            vals = [full_f, full_t, full_v, trim_f, trim_t, trim_v, core_f, core_t, core_v]
            fmts = [_FMT_2DP, _FMT_2DP, _FMT_MONEY,
                    _FMT_2DP, _FMT_2DP, _FMT_MONEY,
                    _FMT_2DP, _FMT_2DP, _FMT_MONEY]
            for ci, (v, fmt) in enumerate(zip(vals, fmts), sc + 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.number_format = fmt
                cell.border = _THIN_BORDER

        r += 1

    return r + 1  # blank gap row


# ---------------------------------------------------------------------------
# 15. Per-hole interval sample table
# ---------------------------------------------------------------------------

def _write_interval_table(
    ws,
    hole_id: str,
    full_samp: pd.DataFrame,
    prices: dict[str, float],
    metal_cutoffs: dict[str, float],
    recovery: float,
    start_row: int,
    start_col: int = _R,
) -> int:
    """Write Full / Trim / Core sample breakdown for one hole. Returns next row."""
    if full_samp.empty:
        return start_row

    sc = start_col
    r  = start_row
    elem_cols = [e for e in prices if e in full_samp.columns]

    trim_samp = _trim_samples(full_samp,    prices, recovery)
    core_samp = _find_core_samples(full_samp, prices, recovery)

    # Hole title banner
    ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc + len(elem_cols) + 3)
    c = ws.cell(row=r, column=sc, value=f"Hole: {hole_id}")
    c.font  = _BOLD_WHITE
    c.fill  = PatternFill("solid", fgColor="404040")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    def _write_variant(variant_name: str, samp: pd.DataFrame, fill):
        nonlocal r
        if samp.empty:
            return

        # Variant header
        frm_v = float(samp["From"].min())
        to_v  = float(samp["To"].max())
        val_v = float(compute_sample_value(samp, prices, recovery).mean())
        title = f"{variant_name}:  {frm_v:.1f}–{to_v:.1f} ft  |  avg ${val_v:.2f}/t"

        ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc + len(elem_cols) + 3)
        th = ws.cell(row=r, column=sc, value=title)
        th.font  = _BOLD
        th.fill  = fill
        th.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 16
        r += 1

        # Column headers
        hdrs = (
            ["From", "To"]
            + [f"{e} (g/t)" if e in _GT_ELEMENTS else f"{e} (%)" for e in elem_cols]
            + ["$/tonne", "Flag"]
        )
        for ci, h in enumerate(hdrs, sc):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font   = _BOLD
            cell.fill   = _BLUE_FILL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        r += 1

        # Sample rows
        flags = compute_sample_flags(samp, metal_cutoffs)
        vals  = compute_sample_value(samp, prices, recovery)
        for idx in samp.index:
            row_s = samp.loc[idx]
            ws.cell(row=r, column=sc,     value=row_s.get("From")).border = _THIN_BORDER
            ws.cell(row=r, column=sc + 1, value=row_s.get("To")).border   = _THIN_BORDER
            for ei, elem in enumerate(elem_cols):
                c2 = ws.cell(row=r, column=sc + 2 + ei, value=row_s.get(elem))
                c2.number_format = _FMT_4DP
                c2.border = _THIN_BORDER
            cv = ws.cell(row=r, column=sc + 2 + len(elem_cols), value=vals.get(idx, 0))
            cv.number_format = _FMT_MONEY
            cv.border = _THIN_BORDER
            cf = ws.cell(row=r, column=sc + 3 + len(elem_cols),
                         value=flags.get(idx, ""))
            cf.alignment = Alignment(horizontal="center")
            cf.border = _THIN_BORDER
            r += 1

    _write_variant("Full",  full_samp,  PatternFill("solid", fgColor="E8F4E8"))
    _write_variant("Trim",  trim_samp,  PatternFill("solid", fgColor="FFF9E6"))
    _write_variant("Core",  core_samp,  PatternFill("solid", fgColor="FFE8E8"))

    return r + 1  # blank gap row


# ---------------------------------------------------------------------------
# 16. Main Excel writer  (BRS010-style single sheet)
# ---------------------------------------------------------------------------

def write_excel(
    cleaned_df: pd.DataFrame,
    intervals: pd.DataFrame,
    prices: dict[str, float],
    cutoff: float,
    output_path: str | Path,
    metal_cutoffs: Optional[dict[str, float]] = None,
    recovery: float = 0.75,
) -> Path:
    """
    Write the BRS010-style single-sheet Excel workbook.

    Layout
    ------
    Left  (cols A–I)  : All sample grades + $/t + flag  (one row per sample)
    Right (cols K+)   : Prices & Cutoffs block
                        Best Intervals summary (one row per hole)
                        Per-hole Full / Trim / Core sample tables
    """
    if not hasattr(output_path, "write"):
        output_path = Path(output_path)
    if metal_cutoffs is None:
        metal_cutoffs = _DEFAULT_METAL_CUTOFFS

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Assay Results"

    elem_cols = [e for e in prices if e in cleaned_df.columns]

    # -----------------------------------------------------------------------
    # LEFT SECTION: all-samples grade table  (cols A–I)
    # -----------------------------------------------------------------------
    L_HOLE, L_FROM, L_TO = 1, 2, 3
    l_elem_start = 4                          # D = first element
    l_val_col    = l_elem_start + len(elem_cols)
    l_flag_col   = l_val_col + 1

    # Header row 1
    for ci, h in enumerate(
        ["Hole ID", "From (ft)", "To (ft)"]
        + [f"{e} (g/t)" if e in _GT_ELEMENTS else f"{e} (%)" for e in elem_cols]
        + ["$/tonne", "Flag"],
        1,
    ):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font   = _BOLD_WHITE
        cell.fill   = _DARK_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Data rows
    sample_flags  = compute_sample_flags(cleaned_df, metal_cutoffs)
    sample_values = compute_sample_value(cleaned_df, prices, recovery)

    for ri, (idx, srow) in enumerate(cleaned_df.iterrows(), start=2):
        ws.cell(row=ri, column=L_HOLE, value=srow.get("Hole_ID")).border = _THIN_BORDER
        ws.cell(row=ri, column=L_FROM, value=srow.get("From")).border    = _THIN_BORDER
        ws.cell(row=ri, column=L_TO,   value=srow.get("To")).border      = _THIN_BORDER
        for ei, elem in enumerate(elem_cols):
            c2 = ws.cell(row=ri, column=l_elem_start + ei, value=srow.get(elem))
            c2.number_format = _FMT_4DP
            c2.border = _THIN_BORDER
        cv = ws.cell(row=ri, column=l_val_col, value=sample_values.get(idx, 0))
        cv.number_format = _FMT_MONEY
        cv.border = _THIN_BORDER
        cf = ws.cell(row=ri, column=l_flag_col, value=sample_flags.get(idx, ""))
        cf.alignment = Alignment(horizontal="center")
        cf.border = _THIN_BORDER

    # Left column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for i in range(len(elem_cols)):
        ws.column_dimensions[get_column_letter(l_elem_start + i)].width = 10
    ws.column_dimensions[get_column_letter(l_val_col)].width   = 12
    ws.column_dimensions[get_column_letter(l_flag_col)].width  = 7

    # Gap column between left and right sections
    gap_col = l_flag_col + 1
    ws.column_dimensions[get_column_letter(gap_col)].width = 4

    # -----------------------------------------------------------------------
    # RIGHT SECTION  (starting at column K = 11)
    # -----------------------------------------------------------------------
    RIGHT_START = max(l_flag_col + 2, 11)    # at least col K

    # -- Prices block --
    next_r = _write_prices_block(
        ws, prices, cutoff, metal_cutoffs, recovery,
        start_row=1, start_col=RIGHT_START,
    )

    # -- Summary block --
    next_r = _write_summary_block(
        ws, cleaned_df, intervals, prices, recovery,
        start_row=next_r, start_col=RIGHT_START,
    )

    # -- Per-hole interval tables --
    for _, irow in intervals.iterrows():
        if bool(irow.get("no_intersection", False)):
            continue
        hole = irow["Hole_ID"]
        mask = (
            (cleaned_df["Hole_ID"] == hole) &
            (pd.to_numeric(cleaned_df["From"], errors="coerce") >= irow["From"]) &
            (pd.to_numeric(cleaned_df["To"],   errors="coerce") <= irow["To"])
        )
        full_samp = cleaned_df[mask].sort_values("From").reset_index(drop=True)
        next_r = _write_interval_table(
            ws, hole, full_samp, prices, metal_cutoffs, recovery,
            start_row=next_r, start_col=RIGHT_START,
        )

    # Right section column widths
    for i in range(12):
        ws.column_dimensions[get_column_letter(RIGHT_START + i)].width = 13

    ws.freeze_panes = ws["A2"]
    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# 17. Main entry point
# ---------------------------------------------------------------------------

def run_pipeline(
    filepath: str | Path,
    prices: dict[str, float],
    cutoff: float,
    output_path: str | Path,
    *,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    sheet: int | str = 0,
    drop_tolerance: float = 0.5,
    min_length_ft: float = 10.0,
    recovery: float = 0.75,
) -> Path:
    """
    Full pipeline: raw Excel -> cleaned data -> intervals -> output Excel.

    Parameters
    ----------
    filepath       : path to input assay Excel file
    prices         : {"Au": 1950, "Ag": 24, "Pb": 0.95, "Zn": 1.20}
                     Precious metals in USD/troy oz, base metals in USD/lb
    cutoff         : AuEq g/t cut-off for qualifying intervals
    output_path    : where to save the output .xlsx
    header_row     : override auto-detection (0-based)
    data_start_row : override auto-detection (0-based)
    sheet              sheet          : sheet index or name in the input file
    drop_tolerance : fraction below cutoff allowed inside an interval
    min_length_ft  : minimum qualifying interval length in feet
    recovery       : metallurgical recovery for base metals (default 0.75)

    Returns
    -------
    Path to the written output file.
    """
    print(f"[1/4] Loading {filepath} ...")
    raw_df = load_raw(filepath, header_row=header_row,
                      data_start_row=data_start_row, sheet=sheet)

    print(f"[2/4] Cleaning data  ({len(raw_df)} rows) ...")
    cleaned_df, _ = clean_dataframe(raw_df)
    elem_found = [e for e in prices if e in cleaned_df.columns]
    print(f"      Elements found: {elem_found}")

    print(f"[3/4] Finding intervals  (cutoff={cutoff} g/t AuEq) ...")
    intervals = find_best_intervals(
        cleaned_df, prices, cutoff,
        drop_tolerance=drop_tolerance,
        min_length_ft=min_length_ft,
        recovery=recovery,
    )
    if len(intervals):
        n_qual = int((~intervals["no_intersection"]).sum())
        print(f"      {n_qual}/{len(intervals)} holes with qualifying intervals")
    else:
        print("      No holes processed.")

    print(f"[4/4] Writing Excel -> {output_path} ...")
    out = write_excel(cleaned_df, intervals, prices, cutoff, output_path,
                      recovery=recovery)
    print("      Done.")
    return out


# ---------------------------------------------------------------------------
# 18. CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    import json

    parser = argparse.ArgumentParser(
        description="Mining Assay Pipeline -- process lab Excel files"
    )
    parser.add_argument("input",  help="Input assay Excel file (.xlsx)")
    parser.add_argument("output", help="Output Excel file (.xlsx)")
    parser.add_argument(
        "--prices",
        default='{"Au": 1950, "Ag": 24, "Pb": 0.95, "Zn": 1.20}',
        help="JSON dict of metal prices e.g. '{\"Au\":1950,\"Ag\":24}'",
    )
    parser.add_argument("--cutoff",     type=float, default=0.5,
                        help="AuEq g/t cutoff (default 0.5)")
    parser.add_argument("--header-row", type=int,   default=None,
                        help="Override header row (0-based)")
    parser.add_argument("--data-start", type=int,   default=None,
                        help="Override data start row (0-based)")
    parser.add_argument("--sheet",      default="0",
                        help="Sheet index or name (default 0)")
    parser.add_argument("--drop-tol",   type=float, default=0.5,
                        help="Drop tolerance fraction (default 0.5)")
    parser.add_argument("--min-length", type=float, default=10.0,
                        help="Minimum interval length in feet (default 10)")
    parser.add_argument("--recovery",   type=float, default=0.75,
                        help="Base metal metallurgical recovery (default 0.75)")

    args = parser.parse_args()
    prices_in = json.loads(args.prices)
    sheet_in  = int(args.sheet) if args.sheet.isdigit() else args.sheet

    run_pipeline(
        filepath       = args.input,
        prices         = prices_in,
        cutoff         = args.cutoff,
        output_path    = args.output,
        header_row     = args.header_row,
        data_start_row = args.data_start,
        sheet          = sheet_in,
        drop_tolerance = args.drop_tol,
        min_length_ft  = args.min_length,
        recovery       = args.recovery,
    )
